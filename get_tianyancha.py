import openpyxl
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import random
 
# 设置 Edge 选项以启动最大化窗口
options = Options()
options.add_argument("--start-maximized")

options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-extensions")
options.add_argument("--disable-popup-blocking")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--remote-debugging-port=9222")


custom_user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
options.add_argument(f'--user-agent={custom_user_agent}')
 
# 获取查询名单
list_name = []
path = r'/Users/zhangjunchao/work/data/获取信用代码数据/企业名称和信用代码.xlsx'
 
wb = openpyxl.load_workbook(path)
wb_sheet = wb['Sheet1']
maxrows = wb_sheet.max_row
for i in range(maxrows - 1):
    name = wb_sheet.cell(i + 2, 1).value
    list_name.append(name)
 
# 初始化 Edge Driver
edge_service = EdgeService(EdgeChromiumDriverManager().install())
driver = webdriver.Edge(service=edge_service, options=options)
url = 'https://www.tianyancha.com/'
driver.get(url)
driver.refresh()

cookies = [{'domain': '.tianyancha.com', 'httpOnly': False, 'name': 'Hm_lpvt_e92c8d65d92d534b0fc290df538b4758', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '1725244951'}, {'domain': '.tianyancha.com', 'expiry': 1727836950, 'httpOnly': False, 'name': 'auth_token', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': 'eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxNTUyMDcyMDIxMiIsImlhdCI6MTcyNTI0NDk1MCwiZXhwIjoxNzI3ODM2OTUwfQ.CZVNDFMewNryuVqHV66sHHAnzs_70RcXgx__OUT2n3TLUe-K4u3SJMACGbm19l4kItn1mpefPBl9RO0WHQMTkg'}, {'domain': '.tianyancha.com', 'expiry': 1727836950, 'httpOnly': False, 'name': 'tyc-user-info-save-time', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '1725244950778'}, {'domain': '.tianyancha.com', 'expiry': 1727836950, 'httpOnly': False, 'name': 'tyc-user-info', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '%7B%22state%22%3A%223%22%2C%22vipManager%22%3A%220%22%2C%22mobile%22%3A%2215520720212%22%2C%22userId%22%3A%223257134%22%2C%22isExpired%22%3A%220%22%7D'}, {'domain': '.tianyancha.com', 'expiry': 1725331332, 'httpOnly': False, 'name': 'bannerFlag', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': 'true'}, {'domain': '.tianyancha.com', 'expiry': 1756780951, 'httpOnly': False, 'name': 'Hm_lvt_e92c8d65d92d534b0fc290df538b4758', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '1725244932'}, {'domain': '.tianyancha.com', 'expiry': 1725292799, 'httpOnly': False, 'name': 'sajssdk_2015_cross_new_user', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '1'}, {'domain': '.tianyancha.com', 'expiry': 1759804951, 'httpOnly': False, 'name': 'sensorsdata2015jssdkcross', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '%7B%22distinct_id%22%3A%223257134%22%2C%22first_id%22%3A%22191b09c805a51f-0db20a88f469598-26001c51-2073600-191b09c805b270%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTkxYjA5YzgwNWE1MWYtMGRiMjBhODhmNDY5NTk4LTI2MDAxYzUxLTIwNzM2MDAtMTkxYjA5YzgwNWIyNzAiLCIkaWRlbnRpdHlfbG9naW5faWQiOiIzMjU3MTM0In0%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%24identity_login_id%22%2C%22value%22%3A%223257134%22%7D%2C%22%24device_id%22%3A%22191b09c805a51f-0db20a88f469598-26001c51-2073600-191b09c805b270%22%7D'}, {'domain': 'www.tianyancha.com', 'httpOnly': False, 'name': 'HWWAFSESID', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '501151f648f012eab23'}, {'domain': '.tianyancha.com', 'expiry': 1759804931, 'httpOnly': False, 'name': 'TYCID', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': 'f493ff3068d411efb58d57dc430694bd'}, {'domain': 'www.tianyancha.com', 'httpOnly': False, 'name': 'csrfToken', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': 'jfcpSkY-t0sg-o6Jm19VDMCt'}, {'domain': '.tianyancha.com', 'expiry': 1759804931, 'httpOnly': False, 'name': 'CUID', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': 'dab1c8b8e045209b86a5834e194692b8'}, {'domain': '.tianyancha.com', 'httpOnly': False, 'name': 'HMACCOUNT', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '33A60AF991EFE04C'}, {'domain': 'www.tianyancha.com', 'httpOnly': False, 'name': 'HWWAFSESTIME', 'path': '/', 'sameSite': 'Lax', 'secure': False, 'value': '1725244927370'}]

for cookie in cookies:
    if 'expiry' in cookie:
        cookie.pop('expiry')  # Selenium 不需要 'expiry' 字段
    # 确保 'secure' 字段与网站协议一致
    if cookie['secure'] and driver.current_url.startswith("https"):
        cookie['secure'] = True
    else:
        cookie['secure'] = False
    driver.add_cookie(cookie)

# 再次获取当前所有 Cookie 并检查是否包含你添加的 Cookie
current_cookies = driver.get_cookies()
added = any(cookie['name'] == 'auth_token' and cookie['value'] == cookies[1]['value'] for cookie in current_cookies)

# 打印结果
print("Cookie added successfully:", added)
print(driver.get_cookies())

 
# 延时，手动扫码登录
sleep(30)
 
cnt = 0
for j in list_name:
    cnt += 1
    driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div/div[3]/div[2]/div[1]/div[1]/input').clear()  # 定位到搜索框
    driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div/div[3]/div[2]/div[1]/div[1]/input').send_keys(j)  # 在搜索框中输入查询企业名单
    try:
        driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div/div[3]/div[2]/div[1]/button').click()
    except:
        driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div/div[3]/div[2]/div[1]/button').click()
    try:
        name_id = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div/div[2]/div/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[3]/div[4]/span').text
    except:
        name_id = "根据商家名称匹配不到数据"
    print(cnt, j, name_id)
    # 写入商家社会信用代码
    wb_sheet.cell(list_name.index(j) + 2, 2, value=name_id)
 
    # 随机暂缓，防止检测到异常
    sleep(random.uniform(0, 3))
    # 跳转回首页，无需再次登录
    try:
        driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/a').click()
    except:
        sleep(10)
        print("errors")
wb.save(path)
wb.close()
driver.close()