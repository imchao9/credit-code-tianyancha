import openpyxl
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import random
 
# 设置 Edge 选项以启动最大化窗口
options = Options()
options.add_argument("--start-maximized")
 
# 获取查询名单
list_name = []
path = r'/Users/zhangjunchao/work/data/获取信用代码数据/企业名称和信用代码.xlsx'
 
wb = openpyxl.load_workbook(path)
wb_sheet = wb['Sheet1']
maxrows = wb_sheet.max_row
for i in range(maxrows - 1):
    name = wb_sheet.cell(i + 2, 1).value
    list_name.append(name)
 
# 初始化 Edge Driver
edge_service = EdgeService(EdgeChromiumDriverManager().install())
driver = webdriver.Edge(service=edge_service, options=options)
url = 'https://www.qcc.com/?utm_source=baidu1&utm_medium=cpc&utm_term=pzsy'
driver.get(url)
driver.refresh()
 
# 延时，手动扫码登录
sleep(30)
 
cnt = 0
for j in list_name:
    cnt += 1
    driver.find_element(By.ID, 'searchKey').clear()  # 定位到搜索框
    driver.find_element(By.ID, 'searchKey').send_keys(j)  # 在搜索框中输入查询企业名单
    try:
        driver.find_element(By.XPATH, '/html/body/div/div[2]/section[1]/div/div/div/div[1]/div/div/span/button').click()
    except:
        driver.find_element(By.XPATH, '/html/body/div/div[2]/section[1]/div/div/div/div[1]/div/div/span/button').click()
    try:
        name_id = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/div[3]/div[1]/span[4]/span/span/span[1]').text
    except:
        name_id = "根据商家名称匹配不到数据"
    print(cnt, j, name_id)
    # 写入商家社会信用代码
    wb_sheet.cell(list_name.index(j) + 2, 2, value=name_id)
 
    # 随机暂缓，防止检测到异常
    sleep(random.uniform(0, 3))
    # 跳转回首页，无需再次登录
    try:
        driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div[1]/nav[1]/div/a[1]').click()
    except:
        print("errors")
wb.save(path)
wb.close()
driver.close()